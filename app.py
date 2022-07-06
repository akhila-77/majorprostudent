from flask import Flask,render_template,request,url_for,redirect,session
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from pyparsing import col
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import hashlib
from csv import writer
import csv
from flask_mail import Mail,Message
from random import randint

def recommend_course(index,datasetname,num_of_rec=5):
    data=pd.read_csv(datasetname,encoding='ISO-8859-1')
    count_vect = CountVectorizer()
    cv_mat = count_vect.fit_transform(data['cname'])
    cosine_sim_mat = cosine_similarity(cv_mat)
    res=pd.DataFrame(data=data['cid'])
    idx =res.index[res['cid'] == index].tolist()[0]
    scores = list(enumerate(cosine_sim_mat[idx]))
    sorted_scores = sorted(scores,key=lambda x:x[1],reverse=True)
    selected_course_indices = [i[0] for i in sorted_scores[0:]]
    selected_course_scores = [i[1] for i in sorted_scores[0:]]
    recommended_result = data['cname'].iloc[selected_course_indices]
    rec_df = pd.DataFrame(recommended_result)
    rec_df['similarity_scores'] = selected_course_scores
    if(datasetname=='course.csv' or datasetname=='laptops.csv'):
        result1 = data['subject'].iloc[selected_course_indices]
        result2 = data['budget'].iloc[selected_course_indices]
        result3 = data['website'].iloc[selected_course_indices]
        result4=data['description'].iloc[selected_course_indices]
        ress=pd.concat([result1,result2,result3,result4],axis=1)
    elif(datasetname=='Placements.csv'):
        result1 = data['cname'].iloc[selected_course_indices]
        result2 = data['Company'].iloc[selected_course_indices]
        result3 = data['salary'].iloc[selected_course_indices]
        result4=data['Experience'].iloc[selected_course_indices]
        result5=data['ApplyLink'].iloc[selected_course_indices]
        result6=data['description'].iloc[selected_course_indices]
        ress=pd.concat([result1,result2,result3,result4,result5,result6],axis=1)
    rec_df = pd.DataFrame(ress)
    return rec_df.head(num_of_rec)
udata=pd.read_csv('unique_courses.csv')
udata=udata.to_numpy()
column_names=[]
app=Flask(__name__)
app.secret_key='project'
mail=Mail(app)
app.config["MAIL_SERVER"]='smtp.gmail.com'
app.config["MAIL_PORT"]=465
app.config["MAIL_USERNAME"]='akhilajulakanti77@gmail.com'
app.config['MAIL_PASSWORD']='123456'                    #you have to give your password of gmail account
app.config['MAIL_USE_TLS']=False
app.config['MAIL_USE_SSL']=True
mail=Mail(app)
otp=randint(000000,999999)
@app.route('/verify',methods=["GET","POST"])
def verify():
    email=request.form["email"]
    session['email']=email
    msg=Message(subject='OTP',sender='akhilajulakanti77@gmail.com',recipients=[email])
    msg.body=str(otp)
    mail.send(msg)
    return render_template('verify.html')
@app.route('/validate',methods=['POST'])
def validate():
    user_otp=request.form['otp']
    if otp==int(user_otp):
        print(session['email'])
        return render_template('updatepassword.html')
    return "<h3>Please Try Again</h3>"
@app.route('/updatepassword',methods=['GET','POST'])
def updatepassword():
    if(request.method=='POST'):
        password1=request.form['password1']
        password2=request.form['password2']
        if(password1==password2):
            wb = load_workbook('signup.xlsx')
            page=wb.active
            for i in range(1, page.max_row+1):
                cell_obj = page.cell(row=i,column=3)
                if(cell_obj.value==session['email']):
                    print(page.cell(row=i,column=2).value)
                    password3=(hashlib.md5(password1.encode())).hexdigest()
                    page.cell(row=i,column=2).value=password3
                    wb.save(filename='signup.xlsx')
            return render_template('login.html',msg="succesfully updated the password")
        else:
            return render_template('updatepassword.html',msg="Please enter the correct password")
    return render_template('updatepassword.html')
@app.route('/forgetpassword')
def forgetpassword():
    return render_template("forgetpassword.html")
@app.route('/')
def home():
    return render_template('home.html')
@app.route('/signup',methods=['GET','POST'])
def signup():
    if request.method == "POST":
        name1=request.form['name']
        password=request.form['password']
        email=request.form['email']
        password1=(hashlib.md5(password.encode())).hexdigest()
        new_row=[name1,password1,email]
        wb = load_workbook('signup.xlsx')
        page=wb.active
        for i in range(1, page.max_row+1):
            cell_obj = page.cell(row=i, column=3)
            if(email==cell_obj.value):
                return render_template("signup.html",msg="User already registerd")
        else:
            page.append(new_row)
            wb.save(filename='signup.xlsx')
            return render_template('signup.html',msg="User Sucesfully registered")
    return render_template('signup.html')
@app.route('/login',methods=['GET','POST'])
def login():
    if request.method == "POST":
        name=request.form['name']
        pwd=request.form['password']
        password=(hashlib.md5(pwd.encode())).hexdigest()
        wb = load_workbook('signup.xlsx')
        page=wb.active
        for i in range(1, page.max_row+1):
            cell_obj = page.cell(row=i,column=1)
            passwordcell=page.cell(row=i,column=2)
            if(cell_obj.value=='admin' and password==passwordcell.value):
                # return render_template('adminhome.html')
                return redirect(url_for("adminhome"))
            elif(name==cell_obj.value and password==passwordcell.value):
                return render_template('index.html',name=name)
            elif(name==cell_obj.value and password!=passwordcell.value):
                return render_template('login.html',msg="Invalid password")
        else:
            return render_template('login.html',msg="User Not Registerd")
    return render_template('login.html')
@app.route('/index')
def index():
    return render_template('index.html')
@app.route('/courses',methods=['GET','POST'])
def career():
    if(request.method=='POST'):
        index=request.form['submit']
        res=recommend_course(int(index),'course.csv')
        print(res)
        res=res.to_numpy()
        column_names=['NAME','BUDGET','WEBSITE','DESCRIPTION']
        return render_template('display.html',r=res,udata=udata[0:7],colnames=column_names)
    return render_template('career.html',udata=udata[0:7])
@app.route('/laptop',methods=['GET','POST'])
def laptop():
    if(request.method=='POST'):
        index=request.form['submit']
        res=recommend_course(int(index),'laptops.csv')
        res=res.to_numpy()
        column_names=['NAME','BUDGET','WEBSITE','DESCRIPTION']
        return render_template('display.html',r=res,udata=udata[8:16],colnames=column_names)
    return render_template('laptop.html',udata=udata[8:16])
@app.route('/feedback',methods=['GET','POST'])
def feedback():
    wb = load_workbook('feedback.xlsx')
    page=wb.active
    data=[]
    for i in range(2, page.max_row+1):
        x=[]
        for j in range(1,7):
            if(j!=3):
                x.append(page.cell(row=i,column=j).value)
        data.append(x)
    if request.method == "POST":
        value=request.form['submit']
        if(value=="submit"):
            name=request.form['name']
            email=request.form['email']
            feedback=request.form['feedback']
            new_row=[page.max_row,name,email,feedback,0,0]
            row=[page.max_row,name,feedback,0,0]
            data.append(row)
            page.append(new_row)
            wb.save(filename='feedback.xlsx')
            return render_template('feedback.html',msg="feedback form submitted successfully",data=data)
        if("dislike" in value):
            x=value.split('+')
            wb = load_workbook('feedback.xlsx')
            page=wb.active
            page.cell(row=(int(x[1])+1),column=6).value=page.cell(row=(int(x[1])+1),column=6).value+1
            wb.save(filename='feedback.xlsx')
            data=[]
            for i in range(2, page.max_row+1):
                x=[]
                for j in range(1,7):
                    if(j!=3):
                        x.append(page.cell(row=i,column=j).value)
                data.append(x)
            return render_template('feedback.html',data=data)
        if("like" in value):
            x=value.split('+')
            wb = load_workbook('feedback.xlsx')
            page=wb.active
            page.cell(row=(int(x[1])+1),column=5).value=page.cell(row=(int(x[1])+1),column=5).value+1
            wb.save(filename='feedback.xlsx')
            data=[]
            for i in range(2, page.max_row+1):
                x=[]
                for j in range(1,7):
                    if(j!=3):
                        x.append(page.cell(row=i,column=j).value)
                data.append(x)
            return render_template('feedback.html',data=data)
            return render_template('feedback.html',data=data)
        
    return render_template('feedback.html',data=data)
@app.route('/placement',methods=['GET','POST'])
def placement():
    if(request.method=='POST'):
        index=request.form['submit']
        res=recommend_course(int(index),'Placements.csv')
        res=res.to_numpy()
        column_names=['NAME','COMPANY','SALARY','EXPERIENCE','APPLY LINK','DESCRIPTION']
        return render_template('display.html',r=res,udata=udata[15:27],colnames=column_names)
    return render_template('placement.html',udata=udata[15:27])
@app.route('/adminhome',methods=['GET','POST'])
def adminhome():
    if(request.method=='POST'):
        value=request.form['submit']
        if(request.form['submit']=="1_1"):
            data=pd.read_csv('course.csv',encoding='ISO-8859-1')
            return render_template('adddetails.html',colnames=data.columns[1:],index="1_1")
        if(request.form['submit']=="2_2"):
            data=pd.read_csv('laptops.csv',encoding='ISO-8859-1')
            return render_template('adddetails.html',colnames=data.columns[1:],index="2_2")
        if(request.form['submit']=="4_4"):
            data=pd.read_csv('Placements.csv',encoding='ISO-8859-1')
            return render_template('adddetails.html',colnames=data.columns[1:],index="4_4")
        if(request.form['submit']=="submit"):
            index=request.form["opt"]
            if(index=="1_1"):
                wb = pd.read_csv('course.csv',encoding='ISO-8859-1')
            elif(index=="2_2"):
                wb=pd.read_csv('laptops.csv',encoding='ISO-8859-1')
            elif(index=="4_4"):
                wb=pd.read_csv('Placements.csv',encoding='ISO-8859-1')
            data=wb.to_numpy()
            colnames=wb.columns
            return render_template('adminhome.html',data=data,colnames=colnames,index=index)
        if('+' in value and 'editing' in value):
            x=value.split('+')
            print(x)
            if(x[1]=="1_1"):
                data=pd.read_csv('course.csv',encoding='ISO-8859-1')
                colnames=data.columns
                ndata=data.to_numpy()
                editdata=ndata[int(x[0])-1]
                return render_template('editform.html',editdata=editdata,index="1_1",colnames=colnames,len=len(colnames))
            if(x[1]=="2_2"):
                data=pd.read_csv('laptops.csv',encoding='ISO-8859-1')
                colnames=data.columns
                ndata=data.to_numpy()
                editdata=ndata[int(x[0])-1]
                return render_template('editform.html',editdata=editdata,index="2_2",colnames=colnames,len=len(colnames))
            if(x[1]=="4_4"):
                data=pd.read_csv('Placements.csv',encoding='ISO-8859-1')
                colnames=data.columns
                ndata=data.to_numpy()
                editdata=ndata[int(x[0])-1]
                return render_template('editform.html',editdata=editdata,index="4_4",colnames=colnames,len=len(colnames))
        if('+' in value and 'delete' in value):
            x=value.split('+')
            cname=""
            if(x[1]=="1_1"):
                cname="course.csv"
            if(x[1]=="2_2"):
                cname="laptops.csv"
            if(x[1]=="3_3"):
                cname="projects.csv"
            if(x[1]=="4_4"):
                cname="Placements.csv"
            updatedlist=[]
            with open(cname,newline="") as f:
                reader=csv.reader(f)
                username=x[0]
                for row in reader: 
                    print(row)
                    if row[0]!=username:
                        updatedlist.append(row)
            with open(cname,"w",newline="") as f:
                Writer=csv.writer(f)
                Writer.writerows(updatedlist)
            data=pd.read_csv(cname)
            for i in range(len(data)):
                data.loc[i,"id"]=i+1
            data.to_csv(cname,index=False)
            return render_template('adminhome.html',msg="THE DETAILS HAVE BEEN REMOVED SUCESSFULLY!")
    return render_template('adminhome.html')
@app.route('/editform',methods=['GET','POST'])
def editform():
    if(request.method=='POST'):
        value=request.form['submit']
        if(value=="1_1"):
            cname="course.csv"
        if(value=="2_2"):
            cname="laptops.csv"
        if(value=="3_3"):
            cname="projects.csv"
        if(value=="4_4"):
            cname="Placements.csv"
        r=[]
        data=pd.read_csv(cname,encoding='ISO-8859-1')
        colnames=data.columns
        for i in range(len(colnames)):
            r.append(request.form[colnames[i]])
        for j in range(len(colnames)):
            data.loc[int(r[0])-1,colnames[j]]=r[j]
        data.to_csv(cname,index=False)
        return render_template('adminhome.html',msg="THE DETAILS HAVE BEEN EDITED SUCESSFULLY!") 
    return render_template('editform.html')
@app.route('/adddetails',methods=['GET','POST'])
def adddetails():
    if(request.method=='POST'):
        res=[]
        index=request.form['submit']
        name=""
        if(index=="1_1"):
            name="course.csv"
        elif(index=="2_2"):
            name="laptops.csv"
        elif(index=="4_4"):
            name="Placements.csv"
        data=pd.read_csv(name,encoding='ISO-8859-1')
        colnames=data.columns
        res.append(len(data)+1)
        for i in range(1,len(colnames)):
            x=request.form[colnames[i]]
            res.append(x)
        with open(name,'a') as f_object:
            writer_object=writer(f_object)
            writer_object.writerow(res)
            f_object.close()  
        return render_template('adminhome.html',msg="THE DETAILS HAVE BEEN ADDED SUCCESSFULLY!")
    return render_template('adddetails.html')
@app.route('/adminfeedback',methods=['GET','POST'])
def adminfeedback():
    wb = load_workbook('feedback.xlsx')
    page=wb.active
    data=[]
    for i in range(2, page.max_row+1):
        id=page.cell(row=i,column=1)
        name= page.cell(row=i,column=2)
        feedback=page.cell(row=i,column=4)
        likes=page.cell(row=i,column=5)
        dislikes=page.cell(row=i,column=6)
        data.append([id.value,name.value,feedback.value,likes.value,dislikes.value])
    if(request.method=='POST'):
        value=request.form["submit"]
        if("delete" in value):
            x=value.split('+')
            print(x)
            wb = load_workbook('feedback.xlsx')
            page=wb.active
            page.delete_rows(int(x[1])+1,1)
            for i in range(2,page.max_row+1):
                page.cell(row=i,column=1).value=i-1
            wb.save(filename='feedback.xlsx')
            return render_template('adminfeedback.html',data=data)
    return render_template('adminfeedback.html',data=data)
if __name__=='__main__':
    app.run(debug=True)
